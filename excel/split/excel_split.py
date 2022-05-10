import logging

from openpyxl import load_workbook, Workbook
from copy import copy
import time
from tqdm import tqdm

# 参数
data_directory = "D:/PycharmProjects/OfficeHighEfficiency/excel/split/src/"  # 保存数据的文件夹，绝对路径
# data_directory = "./src/"  # 相对路径

input_file_path = data_directory + "输入数据.xlsx"  # 输入文件地址
col_to_wbs = 0  # 分工作簿的行标 （从0开始
col_to_sheets = 4  # 分sheet的行标

t1 = time.time()
wb_input = load_workbook(input_file_path)  # 打开工作簿，参数为文件地址
ws_0 = wb_input.active

# 用dict进行分组，两个dict嵌套
group_dict = dict()
for r in ws_0.iter_rows(values_only=True, min_row=2):
    key_0 = r[col_to_wbs]
    key_1 = r[col_to_sheets]
    if key_0 is None or key_1 is None:
        logging.warning("指定列标存在空值，该行已跳过: {}".format(r))
        continue
    group_dict.setdefault(key_0, dict())
    group_dict[key_0].setdefault(key_1, list())
    group_dict[key_0][key_1].append(r)


def set_cell_style(cell_0, cell_1):
    '''
    copy cell_0 to cell_1
    '''
    cell_1.font = copy(cell_0.font)
    cell_1.border = copy(cell_0.border)
    cell_1.fill = copy(cell_0.fill)
    cell_1.number_format = copy(cell_0.number_format)
    cell_1.protection = copy(cell_0.protection)
    cell_1.alignment = copy(cell_0.alignment)


def set_row_style(row_0, row_1):
    for c_idx, c_1 in enumerate(row_1):
        c_0 = row_0[c_idx]
        set_cell_style(c_0, c_1)


for key_0 in group_dict.keys():
    wb = Workbook()  # 为每个key1初始化一个工作簿
    save_path = data_directory + "{}.xlsx".format(key_0)  # 设置保存路径，以key1命名工作簿
    sheet_0 = wb.active  # 默认的（active）sheet，后续要将所有行写入
    sheet_0.append(tuple(ws_0.values)[0])  # 加入标题行

    # 遍历所有key2，将对应行写入对应的sheet
    for key_1 in group_dict[key_0]:
        sheet = wb.create_sheet(title=key_1)  # 在工作簿中创建新的sheet，以key2命名
        sheet.append(tuple(ws_0.values)[0])  # 加入标题行

        for row in tqdm(group_dict[key_0][key_1], desc="写入 文件 - {}, sheet - {}".format(key_0, key_1)):
            sheet.append(row)  # 写入该key2对应的sheet
            sheet_0.append(row)  # 写入默认sheet

    # 64 - 79行修改格式 (耗时，按需求可删除)
    for sheet in wb.worksheets:
        sheet_len = 0
        if sheet.title in group_dict[key_0]:
            sheet_len = len(group_dict[key_0][sheet.title]) + 1  # +1 标题
        else:
            for st in group_dict[key_0].values():
                sheet_len += len(st)
            sheet_len += 1

        for r_idx, r in tqdm(enumerate(sheet.iter_rows()), total=sheet_len,
                             desc="格式调整 文件 - {}, sheet - {}".format(key_0, sheet.title)):
            if r_idx == 0:
                set_row_style(tuple(ws_0.rows)[0], r)
            else:
                set_row_style(tuple(ws_0.rows)[1], r)

    wb.save(filename=save_path)


print("done in {} s".format(round(time.time() - t1, 2)))

    
